import openpyxl

book = openpyxl.load_workbook("C:\\Users\\vicky\\Downloads\\download.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=1, column=2).value = "Rahul"

print(sheet.cell(row=1, column=2).value)
print(sheet.max_row)
print(sheet.max_column)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase 2":

        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value # Rahul

print(Dict)